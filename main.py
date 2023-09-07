from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, NoSuchWindowException
from selenium.webdriver.chrome.service import Service
import openpyxl
import time


def auth(driver, words):
    wait = WebDriverWait(driver, 10)
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[1])
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    driver.get("https://quests.mystenlabs.com/")
    connect_wallet_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//span[contains(text(), 'Connect Wallet')]")))
    connect_wallet_btn.click()
    try:
        check_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="terms"]')))
        check_btn.click()
        accept_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="radix-:Rajm:"]/div[2]/button')))
        accept_btn.click()
    except:
        pass
    # auth
    chose_sui_wallet = wait.until(
        EC.element_to_be_clickable((By.CLASS_NAME, "c-iepcqn")))
    chose_sui_wallet.click()
    wait.until(EC.number_of_windows_to_be(2))
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[1])
    wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "#root > div > div.flex.flex-col.flex-nowrap.items-center.justify-center > div > div > div.flex.sticky.pb-10.m-auto.w-\[300px\].-bottom-px.bg-sui-lightest > a"))).click()
    wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "#root > div > div.flex.flex-col.flex-nowrap.gap-7\.5.mt-7 > div:nth-child(2) > a"))).click()
    wait.until(EC.presence_of_element_located(
        (By.NAME, "mnemonic.0"))).send_keys(words)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.flex.flex-col.flex-nowrap.rounded-20.items-center.bg-sui-lightest.shadow-wallet-content.p-7\.5.pt-10.flex-grow.w-full.max-h-popup-height.max-w-popup-width.overflow-auto > div.mt-7\.5.flex.flex-col.flex-nowrap.items-stretch.flex-1.flex-grow.w-full > form > div > button > div.truncate"))).click()
    wait.until(EC.presence_of_element_located(
        (By.NAME, "password"))).send_keys("7j0AfiiGE1")
    wait.until(EC.presence_of_element_located(
        (By.NAME, "confirmPassword"))).send_keys("7j0AfiiGE1")
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#root > div > div.flex.flex-col.flex-nowrap.rounded-20.items-center.bg-sui-lightest.shadow-wallet-content.p-7\.5.pt-10.flex-grow.w-full.max-h-popup-height.max-w-popup-width.overflow-auto > div.mt-7\.5.flex.flex-col.flex-nowrap.items-stretch.flex-1.flex-grow.w-full > form > div.flex.flex-nowrap.gap-2\.5.mt-5 > button.transition.no-underline.outline-none.group.flex.flex-row.flex-nowrap.items-center.justify-center.gap-2.cursor-pointer.text-body.font-semibold.max-w-full.min-w-0.w-full.bg-hero-dark.text-white.border-none.hover\:bg-hero.focus\:bg-hero.visited\:text-white.active\:text-white\/70.disabled\:bg-hero-darkest.disabled\:text-white.disabled\:opacity-40.h-10.px-5.rounded-xl"))).click()
    try:
        element = wait.until(EC.element_to_be_clickable(
            (By.TAG_NAME, "input")))
        element.send_keys("7j0AfiiGE1")
    except:
        pass
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    chose_sui_wallet = wait.until(
        EC.element_to_be_clickable((By.CLASS_NAME, "c-iepcqn")))
    chose_sui_wallet.click()
    wait.until(EC.number_of_windows_to_be(2))
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[1])
    connect_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="root"]/div/div[1]/div/main/div/div[2]/div/button[2]')))
    connect_btn.click()
    # Проверка на клейм
    wait = WebDriverWait(driver, 10)
    try:
        driver.switch_to.window(driver.window_handles[0])
        claim_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[1]/div/div[1]/div/main/div/div[2]/div/div[2]/button/span")))
        claim_btn.click()
        wait.until(EC.number_of_windows_to_be(2))
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[1])
        approve_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="root"]/div/div[1]/div/main/div/div[2]/div/button[2]')))
        approve_btn.click()
        driver.switch_to.window(driver.window_handles[0])
        congratulations = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.px-3 span')))
        if congratulations.text == "Redeemed":
            print("Успешно!")
            return "Successful"
    except:
        try:
            congratulations = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.px-3 span')))
            if congratulations.text == "Redeemed":
                print("заклеймлено!")
                return "Claimed"
        except:
            return "Failed"


def main():
    wb = openpyxl.load_workbook('bullsharks.xlsx', read_only=True)
    ws = wb.active
    itteration_times = 1
    for row in ws.rows:
        for cell in row:
            words = cell.value
            chromedriver_path = 'chromedriver.exe'
            service = Service(chromedriver_path)
            service.start()
            options = webdriver.ChromeOptions()
            options.add_extension('sui.crx')
            options.add_argument(
                "user-agent=Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36")
            options.add_argument("--ignore-certificate-errors")
            driver = webdriver.Chrome(options=options, service=service)
            driver.maximize_window()
            result = auth(driver, words)
            if result == "Successful":
                wb = openpyxl.load_workbook('bullsharks.xlsx')
                sheet = wb.active
                sheet.cell(row=itteration_times, column=2,
                           value="Successful")
                wb.save('bullsharks.xlsx')
                wb.close()
            elif result == "Claimed":
                wb = openpyxl.load_workbook('bullsharks.xlsx')
                sheet = wb.active
                sheet.cell(row=itteration_times, column=2,
                           value="Claimed")
                wb.save('bullsharks.xlsx')
                wb.close()
            elif result == "Failed":
                wb = openpyxl.load_workbook('bullsharks.xlsx')
                sheet = wb.active
                sheet.cell(row=itteration_times, column=2,
                           value="Failed")
                wb.save('bullsharks.xlsx')
                wb.close()
            itteration_times += 1
    wb.close()


if __name__ == "__main__":
    main()
